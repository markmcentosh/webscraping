from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font

webpage = 'https://coinmarketcap.com/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

wb = xl.Workbook()

ws = wb.active

ws.title = 'Box Office Report'

ws['A1'] = 'Name of Currency'
ws['B1'] = 'Symbol'
ws['C1'] = 'Current Price'
ws['D1'] = '% Change in Last 24 Hrs'
ws['E1'] = 'Corresponding Price Based on % Change'

currency_rows = soup.findAll('tr')

last_price = {"Bitcoin": 0, "Ethereum": 0}

for x in range(1,6):
    td = currency_rows[x].findAll('td')
    name = td[0].text
    symbol = td[1].text
    price = int(td[5].text.replace(",","").replace("$","").replace(".","").replace("%",""))
    percent_change = int(td[7].text.replace(",","").replace("%","").replace("$","").replace(".","").replace("B",""))
    corresponding_price = price*percent_change


    ws['A' + str(x+1)] = name
    ws['B' + str(x+1)] = symbol
    ws['C' + str(x+1)] = "$" + str(price)
    ws['D' + str(x+1)] = str(percent_change) + "%"
    ws['E' + str(x+1)] = "$" + str(corresponding_price) + '%'

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 30


header_font = Font(size=16, bold=True)

for cell in ws[1:1]:
    cell.font = header_font

import keys 
from twilio.rest import Client

current_value = None

wb.save("CryptocurrenciesReport.xlsx")

for x in range[1:6]:
    td = currency_rows[x].findAll("td")
    name = td[2].text
    current_price = int(td[3].text.replace("$", "").replace(",", "").replace(".", ""))
    if name == "Bitcoin" and (current_price - last_price["Bitcoin"]) >= 5:
        message = f"Bitcoin price has changed by ${(current_price - last_price['Bitcoin'])}"
        last_price[name] = current_value
    elif name == "Ethereum" and (current_price - last_price["Ethereum"]) >= 5:
        message = f"Ethereum price has changed by ${(current_price-last_price['Ethereum'])} since last check"
        last_price[name] = current_value
    else:    
        message = "No significant price changes for Bitcoin or Ethereum"

client = Client(keys.accountSID, keys.auth_token)

TwilioNumber = "+2525090001"

mycellphone = "+12149247435"

textmessage = client.messages.create(to=mycellphone, from_=TwilioNumber, body=message)

print("Message sent: ", message.sid)






